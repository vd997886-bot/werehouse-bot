import os
import re
import difflib
from typing import List, Dict, Any, Tuple

from openpyxl import load_workbook

from telegram import Update
from telegram.ext import Application, MessageHandler, CommandHandler, ContextTypes, filters


TOKEN = os.getenv("TOKEN")
EXCEL_FILE = "warehouse.xlsx"  # файл должен лежать в репо рядом с agent.py

# ---- Нормализация (чтобы "ПУ-11", "пу11", "PU 11" искались похоже) ----

_LAT_TO_CYR = str.maketrans({
    "a": "а", "b": "в", "c": "с", "e": "е", "h": "н", "k": "к",
    "m": "м", "n": "п", "o": "о", "p": "р", "t": "т", "x": "х", "y": "у",
    "A": "А", "B": "В", "C": "С", "E": "Е", "H": "Н", "K": "К",
    "M": "М", "N": "П", "O": "О", "P": "Р", "T": "Т", "X": "Х", "Y": "У",
})

def normalize(text: Any) -> str:
    if text is None:
        return ""
    s = str(text).strip()
    s = s.translate(_LAT_TO_CYR)          # латиница похожих букв -> кириллица
    s = s.replace("ё", "е").replace("Ё", "Е")
    s = s.lower()
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"[^0-9a-zа-я\- ]+", " ", s)  # выкидываем мусор
    s = re.sub(r"\s+", " ", s).strip()
    return s

def normalize_key(text: Any) -> str:
    # версия для сравнения: без пробелов и дефисов
    s = normalize(text)
    s = s.replace(" ", "").replace("-", "")
    return s

def yes_no(value: Any) -> str:
    v = normalize(value)
    return "да" if v in ("yes", "y", "true", "1", "да") else "нет"


# ---- Загрузка Excel ----

def load_items() -> List[Dict[str, Any]]:
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active

    items: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        # Ожидаем колонки как у тебя: A.. (номер/название в A)
        # Подстрой если у тебя другой порядок
        number = row[0]
        quantity = row[1] if len(row) > 1 else None
        shelf = row[2] if len(row) > 2 else None
        location = row[3] if len(row) > 3 else None
        passport = row[4] if len(row) > 4 else None
        category = row[5] if len(row) > 5 else None
        serial = row[6] if len(row) > 6 else None
        checked = row[7] if len(row) > 7 else None  # если есть колонка "проверка"

        items.append({
            "number": number,
            "quantity": quantity,
            "shelf": shelf,
            "location": location,
            "passport": passport,
            "category": category,
            "serial": serial,
            "checked": checked,
            "_key": normalize_key(number),
            "_text": normalize(str(number)),
        })

    return items


# ---- Поиск ----

def find_best_matches(query: str, items: List[Dict[str, Any]], limit: int = 5) -> List[Dict[str, Any]]:
    q_norm = normalize(query)
    q_key = normalize_key(query)

    if not q_norm:
        return []

    # 1) Быстрый "частичный" поиск: если запрос входит в название (по ключу без дефисов/пробелов)
    partial = []
    for it in items:
        if q_key and q_key in it["_key"]:
            partial.append(it)

    if partial:
        # сортируем: короче совпадение ближе (пример: "пу" -> "пу-11" выше чем "пусковой...")
        partial.sort(key=lambda x: len(x["_key"]))
        return partial[:limit]

    # 2) Fuzzy: самое похожее по difflib
    scores: List[Tuple[float, Dict[str, Any]]] = []
    for it in items:
        ratio = difflib.SequenceMatcher(None, q_norm, it["_text"]).ratio()
        ratio2 = difflib.SequenceMatcher(None, q_key, it["_key"]).ratio()
        score = max(ratio, ratio2)
        scores.append((score, it))

    scores.sort(key=lambda x: x[0], reverse=True)
    best = [it for (sc, it) in scores if sc >= 0.45]  # порог похожести
    return best[:limit]


def format_item(it: Dict[str, Any]) -> str:
    name = str(it.get("number", ""))
    shelf = it.get("shelf", "")
    loc = it.get("location", "")
    qty = it.get("quantity", "")
    passport = it.get("passport", "")
    category = it.get("category", "")
    serial = it.get("serial", "")
    checked = it.get("checked", "")

    lines = []
    lines.append(f"✅ {name} есть в наличии")
    lines.append(f"📦 Полка: {shelf}, ячейка: {loc}")
    lines.append(f"🔢 Количество: {qty}")
    lines.append(f"📄 Паспорт: {yes_no(passport)}")
    if category not in (None, "", " "):
        lines.append(f"🆕 Категория: {category}")
    if serial not in (None, "", " "):
        lines.append(f"🔑 Серийный номер: {serial}")
    if checked not in (None, "", " "):
        # тут как просила: "да/нет"
        lines.append(f"✔️ Проверка: {yes_no(checked)}")

    return "\n".join(lines)


# ---- Telegram handlers ----

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text(
        "Привет 👋\n"
        "Напиши номер детали или серийный номер — я проверю склад.\n"
        "Можно писать не точно (например: 'пу', 'пу 11', 'ПУ-11')."
    )

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    text = (update.message.text or "").strip()
    if not text:
        return

    try:
        items = load_items()
    except Exception as e:
        await update.message.reply_text(f"Ошибка чтения Excel: {e}")
        return

    matches = find_best_matches(text, items, limit=5)

    if not matches:
        await update.message.reply_text("❌ Не найдено")
        return

    # если совпадений несколько — покажем топ-3 (как “похожие”)
    if len(matches) == 1:
        await update.message.reply_text(format_item(matches[0]))
    else:
        msg = []
        msg.append("Нашла похожие варианты:\n")
        for it in matches[:3]:
            msg.append(format_item(it))
            msg.append("\n" + "—" * 18 + "\n")
        await update.message.reply_text("\n".join(msg).strip())


def main() -> None:
    if not TOKEN:
        raise RuntimeError("TOKEN env var is missing")

    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
